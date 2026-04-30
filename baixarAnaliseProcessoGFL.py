import os
import requests
from datetime import datetime
from flask import Flask, request, send_file
from flask_cors import CORS
from openpyxl import load_workbook

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "https://pedrogeorge.com.br"}})

@app.route("/executar", methods=["POST"])
def executar_macro():
    # 1. Pegar valor enviado pelo HTML
    valor_area = request.form.get("areaMapa")
    if not valor_area:
        return "Área não informada", 400

    # 2. Criar nome do arquivo
    timestamp = datetime.now().strftime("%d%m%Y%H%M%S")
    nome_arquivo = f"Analise_de_Processos_GFL_{timestamp}.xlsm"

    # 3. Usar pasta temporária (IMPORTANTE para servidor)
    caminho_arquivo = os.path.join("/tmp", nome_arquivo)

    # 4. Baixar o arquivo
    url = "https://pedrogeorge.com.br/ExOp/ArquivosExcel/Analise_de_Processos_GFL.xlsm"
    response = requests.get(url)
    if response.status_code != 200:
        return "Erro ao baixar arquivo base", 500

    with open(caminho_arquivo, "wb") as f:
        f.write(response.content)

    # 5. Abrir e manipular Excel (SEM Excel instalado)
    wb = load_workbook(caminho_arquivo, keep_vba=True)
    ws = wb["01-DEN"]

    # Alterar célula F3
    ws["F3"] = valor_area

    # 6. Salvar alterações
    wb.save(caminho_arquivo)

    # 7. Remover arquivo temporário após envio
    @after_this_request
    def remove_file(response):
        try:
            os.remove(caminho_arquivo)
        except Exception as e:
            app.logger.error("Erro ao remover arquivo: %s", e)
        return response

    # 8. Enviar arquivo para download
    return send_file(
        caminho_arquivo,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.ms-excel.sheet.macroEnabled.12"
    )

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=10000)
